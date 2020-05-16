import os
import platform
import subprocess
from math import pi, sin, cos, atan, tan

from openpyxl import load_workbook


class EphemerisBuilder:
    def __init__(self, input_filename):
        self.input_filename = input_filename
        try:
            wb = load_workbook(f'./{self.input_filename}')
            self.wb = wb
        except FileNotFoundError:
            raise FileNotFoundError(f"Please put {self.input_filename} into script directory")

        try:
            # Get sheet names
            sheet_names = self.wb.get_sheet_names()
            self.sheet = self.wb.get_sheet_by_name(sheet_names[0])
        except:
            raise TypeError(f"{self.input_filename} doesn't have any sheets")
        self.tmn = self.smart_input(4)
        self.tmk = self.smart_input(6)
        self.l = self.smart_input(11) / 15
        self.s0 = self.smart_input(19)
        self.h = self.float_from_sheet(7, 2) / 60
        self.n = self.float_from_sheet(16, 2)
        self.d = self.float_from_sheet(17, 2)
        self.alpha = self.smart_input(13) * 15 * pi / 180
        self.d_del = self.smart_input(15) * pi / 180
        self.fi = self.smart_input(9) * pi / 180
        self.s1 = self.star_time(self.tmn, self.n, self.d, self.l, self.s0)  # start time on start of exp
        self.s2 = self.star_time(self.tmk, self.n, self.d, self.l, self.s0)  # start time on end of exp
        if self.s2 > self.s1:
            self.s2 += 24
        if self.tmk < self.tmn:
            self.tmk += 24

        print('connection to sheet established')

    def float_from_sheet(self, r, c):
        try:
            return float(self.sheet.cell(row=r, column=c).value)
        except:
            raise TypeError(f"Cannot convert {self.sheet.cell(row=r, column=c).value} to int")

    def write_to_cell(self, r, c, v):
        self.sheet.cell(row=r, column=c).value = v

    def smart_input(self, n):
        # gets data from sheet in minutes and seconds
        return self.float_from_sheet(n, 2) + self.float_from_sheet(n, 3) / 60

    @staticmethod
    def star_time(tm, n, d, l, s0):
        # calculates start time from decret time
        mu = 9.856 / 3600  # mu const
        mv = tm - n - d
        if mv < 0:
            mv = mv + 24
        m = mv + l
        if m > 24:
            m = m - 24
        sz = m + s0 + (mu * m) - (mu * l)
        if sz < 0:
            sz = sz + 24
        if sz > 24:
            sz = sz - 24
        return sz

    def solve_to_file(self):
        # makes calculations and puts results to file
        f = 5
        s = self.s1
        tm = self.tmn
        i = 1
        while tm <= self.tmk:
            t = s * 15 * pi / 180 - self.alpha
            cosz = sin(self.fi) * sin(self.d_del) + cos(self.fi) * cos(self.d_del) * cos(t)
            z = (atan(-cosz / (-cosz * cosz + 1) ** 0.5) + 2 * atan(1)) * 180 / pi
            A = atan(1 / (sin(self.fi) / tan(t) - tan(self.d_del) * cos(self.fi) / sin(t))) * 180 / pi + 180
            self.write_to_cell(f, 6, self.fix(tm))
            self.write_to_cell(f, 7, round((tm - self.fix(tm)) * 60))
            if tm > 24:
                self.write_to_cell(f, 6, self.fix(tm - 24))
                self.write_to_cell(f, 7, (tm - 24 - self.fix(self.fix(tm - 24))) * 60)
            self.write_to_cell(f, 8, self.fix(z))
            self.write_to_cell(f, 9, round((z - self.fix(z)) * 60))
            self.write_to_cell(f, 10, self.fix(A))
            self.write_to_cell(f, 11, round((A - self.fix(A)) * 60))
            f += 1
            s += self.h
            tm += self.h
            i += 1
            if i > 10000000:
                self.wb.save(filename='output.xlsx')
                return
            self.wb.save(filename='output.xlsx')

    @staticmethod
    def fix(n):
        sign = 1 if n >= 0 else -1
        return sign * int(abs(n))


def open_file(out_file):
    if platform.system() == 'Darwin':  # macOS
        subprocess.call(('open', out_file))
    elif platform.system() == 'Windows':  # Windows
        os.startfile(out_file)
    else:  # linux variants
        subprocess.call(('xdg-open', out_file))


if __name__ == '__main__':
    out_file = 'output.xlsx'
    print('Script.py started')
    builder = EphemerisBuilder("data.xlsm")
    builder.solve_to_file()
    open_file(out_file)
    print('script successfully finished, results in output.xlsm')
