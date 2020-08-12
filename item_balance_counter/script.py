# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook


def get_cell_value(sheet, row, col):
    return sheet.cell(row=row, column=col).value


print('Script.py started')

# Load in the workbook
try:
    leftovers_wb = load_workbook('./ostatki.xlsx')
except:
    print("Please put ostatki.xlsx into script directory")

try:
    # Get sheet names and the first sheet
    sheet_names = leftovers_wb.get_sheet_names()
    left_sheet = leftovers_wb.get_sheet_by_name(sheet_names[0])
except:
    print("ostatki.xlsx has wrong type")

# doing the same thing but for another excel file
# todo: make a function for getting sheet
try:
    # deficit_wb = load_workbook('./Defitsit.xlsx')
    deficit_wb = load_workbook('./ves_tovar.xlsx')
except:
    # print("Please put Defitsit.xlsx into script directory")
    print("Please put ves_tovar.xlsx into script directory")

try:
    sheet_names = deficit_wb.get_sheet_names()
    def_sheet = deficit_wb.get_sheet_by_name(sheet_names[0])
except:
    print("Defitsit.xlsx has wrong type")

leftovers_barcode_col = 5
leftovers_value_range = (7, 14)
current_row = 2
barcode_to_leftovers = {}
while left_sheet.cell(row=current_row, column=1).value is not None or left_sheet.cell(row=current_row,
                                                                                      column=2).value is not None:
    current_row += 1
    barcode = left_sheet.cell(row=current_row, column=5).value
    if barcode is None:
        print(f"skipping barcode {barcode if barcode else 'empty barcode'}")
        continue
    leftovers = 0
    for i in range(leftovers_value_range[0], leftovers_value_range[1]):
        try:
            leftovers += int(left_sheet.cell(row=current_row, column=i).value)
        except:
            pass
    barcode_to_leftovers[barcode] = leftovers

barcodes_to_rows = {}
current_row = 2
while def_sheet.cell(row=current_row, column=9).value is not None or def_sheet.cell(row=current_row,
                                                                                     column=1).value is not None:
    bar = def_sheet.cell(row=current_row, column=9).value
    barcodes_to_rows[bar] = current_row
    current_row += 1

for bar, left in barcode_to_leftovers.items():
    try:
        row = barcodes_to_rows[bar]
    except:
        # print(f"skipping {bar}")
        continue
    def_sheet[f'K{row}'] = left

try:
    deficit_wb.save(filename='ves_tovar.xlsx')
except PermissionError:
    print("Please close the ves_tovar.xlsx before running")
    print("RESULTS ARE NOT SAVED, CLOSE ves_tovar.xlsx AND START AGAIN")

print('script successfully ended, results in ves_tovar.xlsx')
